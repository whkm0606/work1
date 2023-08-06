from dataclasses import fields

from django.core.validators import RegexValidator
from django.core.exceptions import ValidationError
from django.shortcuts import render,redirect
import requests
from django import forms
from django.utils.safestring import mark_safe
from app01.utils.pagintion import Pagination
from app01 import models
from django.views.decorators.csrf import csrf_exempt


def depart_list(request):
    """部门列表"""

    if request.method == 'GET':
        queryset = models.Department.objects.all()
        #create_time.strftime("%Y-%m-%d")
        #obj.get_gender_display()#get_字段名称_display()
        page_object = Pagination(request,queryset,page_size=3)
        context = {
            "depart_list": page_object.page_queryset,  # 分完页的数据
            "page_string": page_object.html  # 生成页码
        }
        return  render(request,'depart_list.html',context)

def depart_add(request):
    """添加部门"""
    if request.method=="GET":
        return  render(request,'depart_add.html')
    title = request.POST.get("title")
    print(title)
    models.Department.objects.create(title = title)
    return redirect("/depart_list/")

def depart_del(request):
    """删除部门"""
    id = request.GET.get("nid")
    models.Department.objects.filter(id = id).delete()
    return redirect('/depart_list/')
def depart_edit(request,nid):
    """编辑部门"""
    if request.method == "GET":
        obj = models.Department.objects.filter(id = nid).first()
        print(obj)
        return render(request, 'depart_edit.html',{"nid": nid,"title":obj.title})
    title = request.POST.get("title")
    print(title,nid)
    models.Department.objects.filter(id=nid).update(title = title)
    return redirect("/depart_list")

@csrf_exempt
def depart_upload(request):
    """批量上传（Excel）文件"""
    # 1  用户上传的文件对象
    fidel_object = request.FILES.get("excel")
    print(fidel_object.name)
    #2 把对象传给openxypl，让它打开Excel读取文件内容
    from openpyxl import load_workbook
    wb = load_workbook(fidel_object)
    sheet = wb.worksheets[0]
    # cell = sheet.cell(1,1)#读取第一行第一列的数据
    # print(cell.value)
    for line in sheet.iter_rows(min_row=2):#从第二列开始
        print(line[0].value)
        text = line[0].value
        object = models.Department.objects
        if not object.filter(title=text).exists():
            object.create(title=text)
    return redirect("/depart_list/")
    # f = open(fidel_object.name, mode='wb')  # 进行上传文件
    # for chunk in fidel_object.chunks():
    #     f.write(chunk)
    # f.close()
    return redirect("/depart_list/")